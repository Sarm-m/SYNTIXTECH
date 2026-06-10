# scripts/datos/generate_import_excel.py

import argparse
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SEED = 20260609

MARCAS = [
    "Chevrolet", "Renault", "Mazda", "Toyota", "Kia",
    "Nissan", "Hyundai", "Ford", "Volkswagen", "Suzuki"
]

MODELOS = [
    "NHR", "Duster", "CX-30", "Hilux", "Sportage",
    "Frontier", "Tucson", "Ranger", "Crafter", "Vitara"
]

TIPOS_VEHICULO = [
    "Camioneta", "Automovil", "Buseta", "Van", "Camion"
]

CATEGORIAS = [
    "B1", "B2", "B3", "C1", "C2", "C3"
]

ASEGURADORAS = [
    "Sura", "Bolivar", "Allianz", "Mapfre", "AXA Colpatria",
    "Seguros Mundial", "Liberty", "Equidad Seguros"
]

CDA_LIST = [
    "CDA Norte", "CDA Sur", "CDA Occidente", "CDA Capital",
    "CDA Movilidad", "CDA Express"
]

RESULTADOS_RTM = [
    "Aprobado", "Rechazado", "Pendiente"
]

VALIDACION_TIPOS = [
    "SOAT", "RTM", "Documentos", "Revision preventiva"
]

VALIDACION_ESTADOS = [
    "Aprobado", "Pendiente", "Rechazado"
]

DISPERSED_EXPIRY_OFFSETS = [
    -120, -60, -30, -15, -7, -1,
    0, 0, 0, 0,
    1, 2, 3, 4, 5, 6, 7, 7,
    8, 10, 12, 15, 18, 22, 26, 30,
    31, 45, 60, 90, 120, 180, 240, 300, 365, 540, 720, 900, 1080, 1200,
]


def make_plate(index: int) -> str:
    """
    Genera placas colombianas simples tipo AAA001.
    Soporta miles de registros sin repetir.
    """
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    a = letters[(index // (26 * 26)) % 26]
    b = letters[(index // 26) % 26]
    c = letters[index % 26]
    number = 100 + (index % 900)
    return f"{a}{b}{c}{number}"


def random_date(start: date, end: date) -> str:
    days = (end - start).days
    selected = start + timedelta(days=random.randint(0, days))
    return selected.isoformat()


def future_date(days_min: int, days_max: int) -> str:
    return (date.today() + timedelta(days=random.randint(days_min, days_max))).isoformat()


def past_date(days_min: int, days_max: int) -> str:
    return (date.today() - timedelta(days=random.randint(days_min, days_max))).isoformat()


def dispersed_expiry(index: int, offset_shift: int = 0) -> date:
    offset = DISPERSED_EXPIRY_OFFSETS[(index + offset_shift) % len(DISPERSED_EXPIRY_OFFSETS)]
    return date.today() + timedelta(days=offset)


def expiry_case(expiry: date) -> str:
    days = (expiry - date.today()).days
    if days < 0:
        return "vencido"
    if days == 0:
        return "vence hoy"
    if days <= 7:
        return "proximo 1-7 dias"
    if days <= 30:
        return "proximo 8-30 dias"
    return "vigente"


def build_conductores(rows: int, dispersed: bool = False):
    data = []

    for i in range(rows):
        documento = str(1000000000 + i)
        nombre = f"Conductor Prueba {i + 1}"
        telefono = f"300{1000000 + i:07d}" if dispersed else f"300{random.randint(1000000, 9999999)}"
        categoria = CATEGORIAS[i % len(CATEGORIAS)] if dispersed else random.choice(CATEGORIAS)
        fecha_vencimiento = dispersed_expiry(i).isoformat() if dispersed else future_date(90, 900)

        data.append({
            "nombre": nombre,
            "documento": documento,
            "telefono": telefono,
            "categoria": categoria,
            "fechaVencimiento": fecha_vencimiento,
        })

    return data


def build_vehiculos(rows: int, conductores):
    data = []

    for i in range(rows):
        placa = make_plate(i)
        marca = random.choice(MARCAS)
        modelo = random.choice(MODELOS)
        anio = random.randint(2015, 2026)
        tipo = random.choice(TIPOS_VEHICULO)
        conductor = conductores[i % len(conductores)]

        data.append({
            "placa": placa,
            "marca": marca,
            "modelo": modelo,
            "anio": anio,
            "tipo": tipo,
            "conductorDocumento": conductor["documento"],
        })

    return data


def build_soats(vehiculos, dispersed: bool = False):
    data = []

    for i, vehiculo in enumerate(vehiculos):
        if dispersed:
            fecha_fin_date = dispersed_expiry(i, 7)
            fecha_inicio_date = fecha_fin_date - timedelta(days=365)
            fecha_expedicion = (fecha_inicio_date - timedelta(days=15)).isoformat()
            fecha_inicio = fecha_inicio_date.isoformat()
            fecha_fin = fecha_fin_date.isoformat()
            observaciones = f"SOAT {expiry_case(fecha_fin_date)} generado para prueba de alerta"
        else:
            fecha_expedicion = past_date(30, 300)
            fecha_inicio = past_date(20, 250)
            fecha_fin = future_date(30, 365)
            observaciones = "Registro generado para prueba masiva de importacion"

        data.append({
            "placa": vehiculo["placa"],
            "numeroPoliza": f"SOAT-{2026}-{i + 1:06d}",
            "aseguradora": random.choice(ASEGURADORAS),
            "fechaExpedicion": fecha_expedicion,
            "fechaInicioVigencia": fecha_inicio,
            "fechaFinVigencia": fecha_fin,
            "observaciones": observaciones,
        })

    return data


def build_rtms(vehiculos, dispersed: bool = False):
    data = []

    for i, vehiculo in enumerate(vehiculos):
        if dispersed:
            fecha_vencimiento_date = dispersed_expiry(i, 14)
            fecha_expedicion = (fecha_vencimiento_date - timedelta(days=365)).isoformat()
            fecha_vencimiento = fecha_vencimiento_date.isoformat()
            observaciones = f"RTM {expiry_case(fecha_vencimiento_date)} generado para prueba de alerta"
        else:
            fecha_expedicion = past_date(30, 300)
            fecha_vencimiento = future_date(30, 365)
            observaciones = "Registro generado para prueba masiva de importacion"

        data.append({
            "placa": vehiculo["placa"],
            "numeroCertificado": f"RTM-{2026}-{i + 1:06d}",
            "cda": random.choice(CDA_LIST),
            "nitCda": f"900{random.randint(100000, 999999)}-{random.randint(0, 9)}",
            "fechaExpedicion": fecha_expedicion,
            "fechaVencimiento": fecha_vencimiento,
            "resultado": RESULTADOS_RTM[i % len(RESULTADOS_RTM)] if dispersed else random.choice(RESULTADOS_RTM),
            "observaciones": observaciones,
        })

    return data


def build_validaciones(vehiculos, rows: int, dispersed: bool = False):
    data = []

    for i in range(rows):
        vehiculo = vehiculos[i % len(vehiculos)]

        tipo = VALIDACION_TIPOS[i % len(VALIDACION_TIPOS)] if dispersed else random.choice(VALIDACION_TIPOS)
        estado = VALIDACION_ESTADOS[i % len(VALIDACION_ESTADOS)] if dispersed else random.choice(VALIDACION_ESTADOS)
        data.append({
            "placa": vehiculo["placa"],
            "tipo": tipo,
            "estado": estado,
            "fecha": (date.today() - timedelta(days=i % 181)).isoformat() if dispersed else past_date(1, 180),
            "observaciones": (
                f"Validacion {estado.lower()} de {tipo} generada para prueba"
                if dispersed
                else f"Validacion automatica de prueba {i + 1}"
            ),
        })

    return data


def build_preferencias():
    return [
        {"clave": "syntix_threshold", "valor": "30"},
        {"clave": "syntix_simulated_date", "valor": ""},
        {"clave": "syntix_dark_mode", "valor": "false"},
    ]


def build_instrucciones():
    return [
        {"tema": "Formato de fechas", "detalle": "Usa YYYY-MM-DD, por ejemplo 2026-05-22."},
        {"tema": "Orden recomendado", "detalle": "Conductores, Vehiculos, SOAT y RTM."},
        {"tema": "Datos de prueba", "detalle": "Archivo generado para validar importaciones masivas."},
    ]


def write_sheet(wb, sheet_name: str, headers, rows):
    ws = wb.create_sheet(title=sheet_name)

    ws.append(headers)

    for item in rows:
        ws.append([item.get(header, "") for header in headers])

    style_sheet(ws, headers)


def style_sheet(ws, headers):
    header_fill = PatternFill("solid", fgColor="10233F")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_index, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_index)
        width = max(14, min(28, len(header) + 4))
        ws.column_dimensions[col_letter].width = width


def generate_excel(rows: int, output_path: str, dispersed: bool = False):
    random.seed(SEED)

    conductores = build_conductores(rows, dispersed)
    vehiculos = build_vehiculos(rows, conductores)
    soats = build_soats(vehiculos, dispersed)
    rtms = build_rtms(vehiculos, dispersed)
    validaciones = build_validaciones(vehiculos, rows, dispersed)
    preferencias = build_preferencias()
    instrucciones = build_instrucciones()

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    write_sheet(
        wb,
        "Instrucciones",
        ["tema", "detalle"],
        instrucciones,
    )

    write_sheet(
        wb,
        "Vehiculos",
        ["placa", "marca", "modelo", "anio", "tipo", "conductorDocumento"],
        vehiculos,
    )

    write_sheet(
        wb,
        "Conductores",
        ["nombre", "documento", "telefono", "categoria", "fechaVencimiento"],
        conductores,
    )

    write_sheet(
        wb,
        "SOAT",
        [
            "placa",
            "numeroPoliza",
            "aseguradora",
            "fechaExpedicion",
            "fechaInicioVigencia",
            "fechaFinVigencia",
            "observaciones",
        ],
        soats,
    )

    write_sheet(
        wb,
        "RTM",
        [
            "placa",
            "numeroCertificado",
            "cda",
            "nitCda",
            "fechaExpedicion",
            "fechaVencimiento",
            "resultado",
            "observaciones",
        ],
        rtms,
    )

    write_sheet(
        wb,
        "Validaciones",
        ["placa", "tipo", "estado", "fecha", "observaciones"],
        validaciones,
    )

    write_sheet(
        wb,
        "Preferencias",
        ["clave", "valor"],
        preferencias,
    )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print("Excel generado correctamente:")
    print(f"Archivo: {output.resolve()}")
    print(f"Conductores: {len(conductores)}")
    print(f"Vehiculos: {len(vehiculos)}")
    print(f"SOAT: {len(soats)}")
    print(f"RTM: {len(rtms)}")
    print(f"Validaciones: {len(validaciones)}")
    print(f"Preferencias: {len(preferencias)}")
    print(f"Instrucciones: {len(instrucciones)}")
    print(f"Modo disperso: {'si' if dispersed else 'no'}")
    print(f"Total aproximado de filas de datos: {len(conductores) + len(vehiculos) + len(soats) + len(rtms) + len(validaciones) + len(preferencias) + len(instrucciones)}")


def main():
    parser = argparse.ArgumentParser(
        description="Genera un Excel de prueba para importar datos masivos en DriveControl/SYNTIXTECH."
    )

    parser.add_argument(
        "--rows",
        type=int,
        default=300,
        help="Cantidad de registros principales a generar por entidad. Default: 300",
    )

    parser.add_argument(
        "--out",
        type=str,
        default="drive-control-import-300.xlsx",
        help="Ruta del archivo Excel de salida.",
    )

    parser.add_argument(
        "--dispersed",
        action="store_true",
        help="Genera vencimientos y estados variados para probar alertas.",
    )

    args = parser.parse_args()

    if args.rows < 250:
        raise ValueError("Usa minimo 250 registros para probar importacion masiva.")

    if args.rows > 10000:
        raise ValueError("No generes mas de 10000 registros por entidad si tu backend tiene ese limite.")

    generate_excel(args.rows, args.out, args.dispersed)


if __name__ == "__main__":
    main()
